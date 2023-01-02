# -*- coding: utf-8 -*-
"""
Created on Fri Nov 25 17:12:22 2022

@author: Lila
"""

import openpyxl
from openpyxl import load_workbook
import time
import os

def write_results(dataset_used, DCAE_output, classification_task,
                  learning_rate, nb_epochs, dropout, accuracy_train,
                  accuracy_test, f1_train, f1_test, 
                  excel_path, excel_name = "model_runs"):
  """
  This function writes information given in input in an excel file.

  Inputs:
  - dataset_used (str): name of the dataset used for train and test.
  - DCAE_output (str): None if raw data, name of DCAE architecture if data
    is from DCAE output.
  - classification_task	(str): classification task 
  - learning_rate (float): model's learning rate 
  - nb_epochs (int): model's number of epochs  
  - dropout (float): model's dropout
  - accuracy_train,accuracy_test (float): accuracy obtained on train and test sets.
  - f1_train,f1_test (float): f1 score obtained on train and test sets
  - excel_path (str): path of the excel file
  - excel_name (str): name of the excel file (without extension .xlsx)
  """

  # Open file and initalise variables
  # =======================================
  #----- Open xlsx file
  os.chdir(excel_path) #go to correct folder 
  wb = load_workbook(filename = excel_path + "/" + excel_name + ".xlsx")
  #----- Get the current Active Sheet
  ws = wb.active #or wb['SHEET_NAME']
  #----- Specify row index
  row = 2
  #----- If there is no first run (true only at the first use) 
  if not(ws.cell(row,1).value): 
    ws.cell(row,1).value = 1 
  else: #increase number of runs 
    ws.cell(row,1).value = ws.cell(row,1).value + 1 
  #----- Update row index according to the number of runs
  row = int(ws.cell(row,1).value) + 2

  # Write information
  # =======================================
  date = time.strftime("%Y-%m-%d %H:%M:%S")
  ws.cell(row,2).value  = date 
  ws.cell(row,3).value  = dataset_used 
  ws.cell(row,4).value  = DCAE_output 
  ws.cell(row,5).value  = classification_task
  ws.cell(row,6).value  = str(learning_rate)
  ws.cell(row,7).value  = str(nb_epochs)
  ws.cell(row,8).value  = str(dropout)
  ws.cell(row,9).value  = str(round(accuracy_train,3))
  ws.cell(row,10).value = str(round(f1_train,3))
  ws.cell(row,11).value = str(round(accuracy_test,3))
  ws.cell(row,12).value = str(round(f1_test,3))
    
  # WRITE now all changes into the file 
  # =======================================
  wb.save(excel_path + "/" + excel_name + ".xlsx")